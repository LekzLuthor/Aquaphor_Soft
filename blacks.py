import os
import pprint
import datetime
import json

import openpyxl


def load_database():
    with open("sours/settings.json", "r") as file:  # достаёт настройки из json файла
        settings = json.load(file)
    files_name = os.listdir(settings['pathway'])

    for f_index, f in enumerate(files_name):

        excel_file = openpyxl.open(f'{settings["pathway"]}/{f}', read_only=True)
        sheet = excel_file.active

        start_line_ind = 0
        while sheet[f'B{start_line_ind}'].value != "№ п/п":
            start_line_ind += 1
        start_line_ind += 3

        end_line_ind = start_line_ind + 1
        while sheet[f'B{end_line_ind}'].value is not None:
            end_line_ind += 1

        equipment = []
        for ind in range(start_line_ind, end_line_ind + 1):
            line = [i.value for i in sheet[f'A{ind}':f'L{ind}'][0]]
            equipment.append(line)
        print(equipment)


def create_report():
    database = {'0': [['Отдел контроля качества',
                       1,
                       'Анализатор влажности',
                       'AND MX-50',
                       None,
                       None,
                       'P1060560',
                       '1 год',
                       datetime.datetime(2021, 7, 28, 0, 0),
                       datetime.datetime(2022, 7, 28, 0, 0),
                       None,
                       '..\\Метрологическая документация 21\\Анализатор влажности AND MX-50 '
                       'Зав№ P1060560 (отдел контроля качества).asice'],
                      ['Отдел контроля качества',
                       1,
                       'Анализатор влажности',
                       'Kern MLB-50-3C',
                       '0-50 г            0,001 г',
                       None,
                       'WL141587',
                       '1 год',
                       datetime.datetime(2020, 7, 8, 0, 0),
                       datetime.datetime(2021, 7, 8, 0, 0),
                       None,
                       '..\\Метрологическая документация 20\\Анализатор влажности Kern '
                       'MLB-50-3C Зав. №WL141587 (СКК).asice'],
                      ['Отдел контроля качества',
                       2,
                       'Весы электронные',
                       'CAS BW-6R',
                       '0-6000 г               2 г',
                       None,
                       '0111В00329',
                       '1 год',
                       datetime.datetime(2021, 7, 13, 0, 0),
                       datetime.datetime(2022, 7, 13, 0, 0),
                       None,
                       '..\\Метрологическая документация 21\\Весы электронные CAS BW-6R Зав.№ '
                       '0111B00329 (отдел контроля качества).asice'],
                      ['Отдел контроля качества',
                       3,
                       'Глубиномер цифровой',
                       'Vogel',
                       '0,01 мм               0-500 мм',
                       None,
                       'С1109260080',
                       '1 год',
                       datetime.datetime(2022, 1, 4, 0, 0),
                       datetime.datetime(2023, 1, 4, 0, 0),
                       None,
                       'Глубиномер цифровой Vogel Зав.№ C1109260080(СКК).asice'],
                      ['Отдел контроля качества',
                       4,
                       'Динамометрический ключ с индикатором часового типа',
                       'DB50N-S TOHNICHI',
                       None,
                       None,
                       '311725Е',
                       '1 год',
                       datetime.datetime(2022, 2, 14, 0, 0),
                       datetime.datetime(2023, 2, 14, 0, 0),
                       None,
                       'Динамометрический ключ с индикатором часового типа DB50N-S TOHNICHI '
                       'Зав.№ 311725E.asice'],
                      ['Отдел контроля качества',
                       5,
                       'Калибр-кольцо',
                       'G1/2, ПР/НЕПР',
                       None,
                       None,
                       '82150027/82150127',
                       '5 лет',
                       datetime.datetime(2020, 10, 2, 0, 0),
                       datetime.datetime(2025, 10, 2, 0, 0),
                       None,
                       '..\\Метрологическая документация 20\\Калибр-кольцо G12, ПР НЕПР Зав. '
                       '№821500278 2150127 (Служба качества).asice']]
                }
    for list_num in database.keys():
        for eq_num in range(len(database[list_num])):
            if database[list_num][eq_num][9].date() < datetime.date.today():
                print(f"{database[list_num][eq_num][9].date()} - Просрочено")
            else:
                print(f"{database[list_num][eq_num][9].date()} - Дата проверки не подошла")


# def february_29_erorr():
#     day, month, year = int(self.database[list_num][eq_num][9].split('.')[0]), int(
#         self.database[list_num][eq_num][9].split('.')[1]), int(
#         self.database[list_num][eq_num][9].split('.')[2])
#
#     print(f'day - {day}, month - {month}, year - {year}')
#     # date = datetime.datetime(year, month, day)
#     # if date < datetime.date.today():
#     #     print(date, '- Просрочено (FIXED)')
#     # else:
#     #     print(date, '- Дата калибровки не просрочена (FIXED)')


def report_to_excel():
    report = {'1': [['Отдел контроля качества',
                     1,
                     'Анализатор влажности',
                     'AND MX-50',
                     None,
                     None,
                     'P1060560',
                     '1 год',
                     datetime.datetime(2021, 7, 28, 0, 0),
                     datetime.datetime(2022, 7, 28, 0, 0),
                     None,
                     '..\\Метрологическая документация 21\\Анализатор влажности AND MX-50 '
                     'Зав№ P1060560 (отдел контроля качества).asice'],
                    ['Отдел контроля качества',
                     1,
                     'Анализатор влажности',
                     'Kern MLB-50-3C',
                     '0-50 г            0,001 г',
                     None,
                     'WL141587',
                     '1 год',
                     datetime.datetime(2020, 7, 8, 0, 0),
                     datetime.datetime(2021, 7, 8, 0, 0),
                     None,
                     '..\\Метрологическая документация 20\\Анализатор влажности Kern '
                     'MLB-50-3C Зав. №WL141587 (СКК).asice'],
                    ['Отдел контроля качества',
                     2,
                     'Весы электронные',
                     'CAS BW-6R',
                     '0-6000 г               2 г',
                     None,
                     '0111В00329',
                     '1 год',
                     datetime.datetime(2021, 7, 13, 0, 0),
                     datetime.datetime(2022, 7, 13, 0, 0),
                     None,
                     '..\\Метрологическая документация 21\\Весы электронные CAS BW-6R Зав.№ '
                     '0111B00329 (отдел контроля качества).asice'],
                    ['Отдел контроля качества',
                     3,
                     'Глубиномер цифровой',
                     'Vogel',
                     '0,01 мм               0-500 мм',
                     None,
                     'С1109260080',
                     '1 год',
                     datetime.datetime(2022, 1, 4, 0, 0),
                     datetime.datetime(2023, 1, 4, 0, 0),
                     None,
                     'Глубиномер цифровой Vogel Зав.№ C1109260080(СКК).asice'],
                    ['Отдел контроля качества',
                     7,
                     'Кронциркуль',
                     'Insize',
                     '0-20 мм         0,001 мм',
                     None,
                     '1608С2065',
                     '1 год',
                     datetime.datetime(2021, 9, 14, 0, 0),
                     datetime.datetime(2022, 9, 14, 0, 0),
                     None,
                     '..\\Метрологическая документация 21\\Кронциркуль Insize Зав.№ '
                     '1608C2065 (отдел контроля качества).asice'],
                    ['Отдел контроля качества',
                     8,
                     'Манометр гидрозаполненный',
                     'Wika',
                     None,
                     datetime.datetime(2014, 11, 1, 0, 0),
                     'БМ 1/Н',
                     '1 год',
                     datetime.datetime(2019, 2, 2, 0, 0),
                     datetime.datetime(2020, 2, 1, 0, 0),
                     None,
                     None],
                    ['Отдел контроля качества',
                     9,
                     'Манометр гидрозаполненный',
                     'Wika',
                     None,
                     datetime.datetime(2014, 11, 1, 0, 0),
                     'БМ 2/Н',
                     '1 год',
                     datetime.datetime(2019, 2, 2, 0, 0),
                     datetime.datetime(2020, 2, 1, 0, 0),
                     None,
                     None],
                    ['Отдел контроля качества',
                     10,
                     'Манометр гидрозаполненный',
                     'Wika',
                     None,
                     datetime.datetime(2014, 11, 1, 0, 0),
                     'БМ 3/Н',
                     '1 год',
                     datetime.datetime(2019, 2, 2, 0, 0),
                     datetime.datetime(2020, 2, 1, 0, 0),
                     None,
                     None]]
              }
    book = openpyxl.Workbook()
    sheet = book.active
    row = 1
    col = 1
    for key in report.keys():
        sheet.cell(row=row, column=col).value = 'filename'
        row += 1
        for equip in report[key]:
            for item in equip:
                sheet.cell(row=row, column=col).value = item
                col += 1
            col = 1
            row += 1

    book.save("report.xlsx")
    book.close()


report_to_excel()

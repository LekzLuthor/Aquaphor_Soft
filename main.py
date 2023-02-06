# дополнительные библиотеки
import json
import sys
import datetime
import os

# для создания xlsx файлов
import openpyxl

# pyqt библиотеки (для интерфейса)
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt5 import uic

# библиотеки для отправки на почту
import smtplib
from email.message import EmailMessage


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('main_window.ui', self)
        self.progressBar.setValue(0)
        self.setWindowTitle("отчеты калибровки Aquaphor International")

        # блок данных программы
        self.settings = {}
        self.excel_files_names = []  # список названий файлов
        self.database = {}  # база данных сформированная из файлов
        self.equipment_report = {}  # оборудование с просроченной датой калибровки
        self.emails = []

        # данные почты для отправки писем
        self._SENDER = 'calibration.aquaphor@gmail.com'
        self._SENDER_PASSWORD = 'ogwgkvtqnvjsfljr'

        # привязка кнопок + редактура полей
        self.generateReportButton.clicked.connect(self.final_build)
        self.saveChangesButton.clicked.connect(self.save_changes)
        self.emailSetForm.setPlaceholderText("example@aquaphor.com")
        self.pathwaySetForm.setPlaceholderText("C:/path/path with your Excel Files")

        # блок логики программы при запуске
        with open("sours/settings.json", "r") as file:  # достаёт настройки из json файла
            self.settings = json.load(file)
            for email in self.settings["emails"]:
                self.emails.append(email)

    def save_changes(self):
        self.generateReportButton.setEnabled(False)
        self.saveChangesButton.setEnabled(False)
        email = self.emailSetForm.toPlainText()
        if email != "":
            if self.check_email(email):
                if email not in self.settings["emails"]:
                    self.settings["emails"].append(email)
                    self.emailStatusBar.setText('mail added to list')
                else:
                    self.emailStatusBar.setText("you haven't entered a new email")
            else:
                pass

        # Получение файлового пути и кэтч ошибки с неправильным (\) символом
        if self.pathwaySetForm.toPlainText().strip():
            try:
                pathway = self.pathwaySetForm.toPlainText()
                try:
                    with open(f'{pathway}/Your Excel Files Will Be here', 'w') as f:
                        f.writelines('Checking the correctness of path way')
                    if os.path.isfile(f'{pathway}/Your Excel Files Will Be here'):
                        os.remove(f'{pathway}/Your Excel Files Will Be here')

                except OSError:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Error: WRONG PATH WAY")
                    msg.setWindowTitle("Warning")
                    retval = msg.exec_()
                    self.pathwayStatusBar.setText('wrong path way')

                if pathway != '':
                    self.settings["pathway"] = pathway
                    self.pathwayStatusBar.setText('pathway saved')

            except SyntaxError:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Error: WRONG PATH WAY")
                msg.setWindowTitle("Warning")
                retval = msg.exec_()
                self.pathwayStatusBar.setText('wrong path way')

        with open("sours/settings.json", "w") as file:  # сохраняет настройки в json файл
            json.dump(self.settings, file)

        with open("sours/settings.json", "r") as file:  # достаёт настройки из json файла
            self.settings = json.load(file)
            for email in self.settings["emails"]:
                self.emails.append(email)

        self.generateReportButton.setEnabled(True)
        self.saveChangesButton.setEnabled(True)

    @staticmethod
    def check_email(email):  # проверка правильности ввода почты
        if email and (
                "@aquaphor.com" in email or "@mail.ru" in email or "@gmail.com" in email or "@yandex.com" in email or
                "@mail.com" in email or "@yandex.com" in email or "@yandex.ru" in email or "@rambler.com" in email or
                "@gmail.ru" in email or "@y.com" in email
        ):
            return True
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Error: wrong email")
            msg.setWindowTitle("Warning")
            retval = msg.exec_()
            return False

    def check_pathway(self):
        if "pathway" in self.settings.keys():
            return True
        return False

    def load_database(self):
        files_name = os.listdir(self.settings['pathway'])
        files_name = [file for file in files_name if file.endswith('xlsx')]
        self.excel_files_names = files_name
        for f_index, f in enumerate(files_name, 1):
            excel_file = openpyxl.open(f'{self.settings["pathway"]}/{f}', read_only=True)
            sheet = excel_file.active

            start_line_ind = 0
            while sheet[f'B{start_line_ind}'].value != "№ п/п":
                start_line_ind += 1
            start_line_ind += 3

            end_line_ind = start_line_ind + 1
            while sheet[f'B{end_line_ind}'].value is not None:
                end_line_ind += 1

            equipment = []
            for ind in range(start_line_ind, end_line_ind + 1):
                line = [i.value for i in sheet[f'A{ind}':f'L{ind}'][0]]
                equipment.append(line)

            self.database[str(f_index)] = equipment

    def create_report(self):
        for list_num in self.database.keys():
            equipment_to_report = []
            for eq_num in range(len(self.database[list_num])):
                if self.database[list_num][eq_num][9]:
                    try:
                        date_delta = datetime.date.today() - self.database[list_num][eq_num][9].date()
                        # добавление если срок калибровки уже истёк
                        if self.database[list_num][eq_num][9].date() < datetime.date.today():
                            new_eq_format = [
                                self.database[list_num][eq_num][0], self.database[list_num][eq_num][2],
                                self.database[list_num][eq_num][3], self.database[list_num][eq_num][4],
                                self.database[list_num][eq_num][6], self.database[list_num][eq_num][7],
                                self.database[list_num][eq_num][8].date(), self.database[list_num][eq_num][9].date(),
                                self.database[list_num][eq_num][10]
                            ]
                            equipment_to_report.append(new_eq_format)
                            # 0 2 3 4 6 7 8 9 10 11
                        # добавление если до срока калибровки осталось меньше месяца
                        elif date_delta.days > -30:
                            new_eq_format = [
                                self.database[list_num][eq_num][0], self.database[list_num][eq_num][2],
                                self.database[list_num][eq_num][3], self.database[list_num][eq_num][4],
                                self.database[list_num][eq_num][6], self.database[list_num][eq_num][7],
                                self.database[list_num][eq_num][8].date(), self.database[list_num][eq_num][9].date(),
                                self.database[list_num][eq_num][10], '',
                                f'до калибровки осталось {abs(date_delta.days)} дней'
                            ]
                            equipment_to_report.append(new_eq_format)
                    except Exception:
                        print('29 FEBRUARY ERROR')
            self.equipment_report[list_num] = equipment_to_report

    def report_to_excel(self):
        if self.check_pathway():
            self.progressBar.setValue(20)
            self.load_database()
            self.progressBar.setValue(40)
            self.create_report()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("You didn't specify path way")
            msg.setWindowTitle("Warning")
            retval = msg.exec_()
            return
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        col = 1
        file_name_index = 0
        sheet.cell(row=1, column=1).value = "Подразделение, использующее ИО, участок"
        sheet.cell(row=1, column=2).value = "Наименование ИО"
        sheet.cell(row=1, column=3).value = "Тип/вид/марка/модель ИО, НД"
        sheet.cell(row=1, column=4).value = "Лимит и точность измерений "
        sheet.cell(row=1, column=5).value = "Заводской (инв.) №"
        sheet.cell(row=1, column=6).value = "Межповерочный интервал"
        sheet.cell(row=1, column=7).value = "Дата очередных поверок (калибровок)"
        sheet.cell(row=1, column=9).value = "Примечание"
        for key in self.equipment_report.keys():
            row += 3
            sheet.cell(row=row, column=col).value = self.excel_files_names[file_name_index]
            row += 2
            for equip in self.equipment_report[key]:
                for item in equip:
                    sheet.cell(row=row, column=col).value = item
                    col += 1
                col = 1
                row += 1
        book.save("sours/reports/calibration report.xlsx")
        book.close()

    def send_to_email(self):
        for email in self.emails:
            msg = EmailMessage()
            msg['Subject'] = f'{datetime.date.today()} Report'
            msg['From'] = self._SENDER
            msg['To'] = email

            with open("sours/reports/calibration report.xlsx", 'rb') as f:
                file_data = f.read()
            msg.add_attachment(file_data, maintype="application", subtype="xlsx", filename="calibration report.xlsx")

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(self._SENDER, self._SENDER_PASSWORD)
                smtp.send_message(msg)

    def final_build(self):
        self.progressBar.setValue(0)
        self.report_to_excel()
        self.progressBar.setValue(60)
        self.send_to_email()
        self.progressBar.setValue(100)
        print('done')


def excepthook(exc_type, exc_value, exc_tb):
    tb = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print("Oбнаружена ошибка !:", tb)


sys.excepthook = excepthook

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
